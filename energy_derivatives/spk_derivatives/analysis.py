"""
Sensitivity Analysis for Energy Derivatives
============================================

Convenience utilities for generating sensitivity tables and stress test scenarios.

Key Functions:
-----------
sensitivity_table(): Greeks at different spot prices
stress_test_volatility(): Price across volatility range
stress_test_rates(): Price across interest rate range
combined_stress_test(): 2D volatility × rate scenarios
export_to_excel(): Save results to Excel workbook
"""

import numpy as np
import pandas as pd
from typing import Dict, List, Optional, Tuple
from .binomial import BinomialTree
from .monte_carlo import MonteCarloSimulator
from .sensitivities import GreeksCalculator


def sensitivity_table(S0: float, K: float, T: float, r: float, sigma: float,
                      spot_range: Optional[List[float]] = None,
                      method: str = 'binomial', N: int = 100,
                      num_simulations: int = 10000) -> pd.DataFrame:
    """
    Generate sensitivity table: Greeks at different spot prices.
    
    Parameters
    ----------
    S0 : float
        Initial spot price
    K : float
        Strike price
    T : float
        Time to maturity
    r : float
        Risk-free rate
    sigma : float
        Volatility
    spot_range : List[float], optional
        Range of spot prices to evaluate (default: -20% to +20% of S0)
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Steps for binomial (default: 100)
    num_simulations : int
        Paths for Monte Carlo (default: 10000)
    
    Returns
    -------
    pd.DataFrame
        Columns: Spot Price, Option Price, Delta, Gamma, Vega, Theta, Rho
    
    Example
    -------
    >>> table = sensitivity_table(S0=0.035, K=0.040, T=1.0, r=0.025, sigma=0.42)
    >>> print(table)
    """
    if spot_range is None:
        # Default: -20% to +20% in 11 steps
        spot_range = np.linspace(S0 * 0.8, S0 * 1.2, 11)
    else:
        spot_range = np.array(spot_range)
    
    results = []
    
    for spot in spot_range:
        # Price
        if method == 'binomial':
            tree = BinomialTree(S0=spot, K=K, T=T, r=r, sigma=sigma, N=N, payoff_type='call')
            price = tree.price()
        else:  # monte_carlo
            mc = MonteCarloSimulator(S0=spot, K=K, T=T, r=r, sigma=sigma, 
                                    num_simulations=num_simulations, payoff_type='call')
            price = mc.price()
        
        # Greeks
        greeks = GreeksCalculator(S0=spot, K=K, T=T, r=r, sigma=sigma, pricing_method=method)
        delta = greeks.delta()
        gamma = greeks.gamma()
        vega = greeks.vega()
        theta = greeks.theta()
        rho = greeks.rho()
        
        results.append({
            'Spot Price ($/kWh)': spot,
            'Option Price ($/kWh)': price,
            'Delta': delta,
            'Gamma': gamma,
            'Vega': vega,
            'Theta': theta,
            'Rho': rho,
        })
    
    df = pd.DataFrame(results)
    return df


def stress_test_volatility(S0: float, K: float, T: float, r: float,
                           vol_range: Optional[List[float]] = None,
                           method: str = 'binomial', N: int = 100,
                           payoff_type: str = 'call') -> pd.DataFrame:
    """
    Stress test: Option price across volatility range.
    
    Parameters
    ----------
    S0 : float
        Spot price
    K : float
        Strike price
    T : float
        Time to maturity
    r : float
        Risk-free rate
    vol_range : List[float], optional
        Volatility levels to test (default: 10% to 100% in 10% steps)
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Binomial steps (default: 100)
    payoff_type : str
        'call' or 'put'
    
    Returns
    -------
    pd.DataFrame
        Columns: Volatility, Option Price, Delta, Vega
    
    Example
    -------
    >>> stress = stress_test_volatility(S0=0.035, K=0.040, T=1.0, r=0.025)
    >>> print(stress)
    """
    if vol_range is None:
        vol_range = np.arange(0.10, 1.01, 0.10)
    else:
        vol_range = np.array(vol_range)
    
    results = []
    
    for vol in vol_range:
        if method == 'binomial':
            tree = BinomialTree(S0=S0, K=K, T=T, r=r, sigma=vol, N=N, payoff_type=payoff_type)
            price = tree.price()
        else:
            mc = MonteCarloSimulator(S0=S0, K=K, T=T, r=r, sigma=vol, payoff_type=payoff_type)
            price = mc.price()
        
        greeks = GreeksCalculator(S0=S0, K=K, T=T, r=r, sigma=vol, pricing_method=method)
        delta = greeks.delta()
        vega = greeks.vega()
        
        results.append({
            'Volatility': f"{vol:.1%}",
            'Volatility (decimal)': vol,
            'Option Price ($/kWh)': price,
            'Delta': delta,
            'Vega': vega,
        })
    
    df = pd.DataFrame(results)
    return df[['Volatility', 'Option Price ($/kWh)', 'Delta', 'Vega']]


def stress_test_rates(S0: float, K: float, T: float, sigma: float,
                      rate_range: Optional[List[float]] = None,
                      method: str = 'binomial', N: int = 100,
                      payoff_type: str = 'call') -> pd.DataFrame:
    """
    Stress test: Option price across interest rate range.
    
    Parameters
    ----------
    S0 : float
        Spot price
    K : float
        Strike price
    T : float
        Time to maturity
    sigma : float
        Volatility
    rate_range : List[float], optional
        Interest rates to test (default: 0% to 10% in 1% steps)
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Binomial steps (default: 100)
    payoff_type : str
        'call' or 'put'
    
    Returns
    -------
    pd.DataFrame
        Columns: Rate, Option Price, Rho
    
    Example
    -------
    >>> stress = stress_test_rates(S0=0.035, K=0.040, T=1.0, sigma=0.42)
    >>> print(stress)
    """
    if rate_range is None:
        rate_range = np.arange(0.00, 0.11, 0.01)
    else:
        rate_range = np.array(rate_range)
    
    results = []
    
    for rate in rate_range:
        if method == 'binomial':
            tree = BinomialTree(S0=S0, K=K, T=T, r=rate, sigma=sigma, N=N, payoff_type=payoff_type)
            price = tree.price()
        else:
            mc = MonteCarloSimulator(S0=S0, K=K, T=T, r=rate, sigma=sigma, payoff_type=payoff_type)
            price = mc.price()
        
        greeks = GreeksCalculator(S0=S0, K=K, T=T, r=rate, sigma=sigma, pricing_method=method)
        rho = greeks.rho()
        
        results.append({
            'Rate': f"{rate:.1%}",
            'Rate (decimal)': rate,
            'Option Price ($/kWh)': price,
            'Rho': rho,
        })
    
    df = pd.DataFrame(results)
    return df[['Rate', 'Option Price ($/kWh)', 'Rho']]


def combined_stress_test(S0: float, K: float, T: float,
                         vol_range: Optional[List[float]] = None,
                         rate_range: Optional[List[float]] = None,
                         method: str = 'binomial', N: int = 100,
                         payoff_type: str = 'call') -> pd.DataFrame:
    """
    2D stress test: Option price at volatility × rate combinations.
    
    Parameters
    ----------
    S0 : float
        Spot price
    K : float
        Strike price
    T : float
        Time to maturity
    vol_range : List[float], optional
        Volatilities to test (default: 20%, 40%, 60%)
    rate_range : List[float], optional
        Rates to test (default: 0%, 2.5%, 5%)
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Binomial steps (default: 100)
    payoff_type : str
        'call' or 'put'
    
    Returns
    -------
    pd.DataFrame
        Index: Volatility, Columns: Rate values, Values: Option prices
    
    Example
    -------
    >>> stress = combined_stress_test(S0=0.035, K=0.040, T=1.0)
    >>> print(stress)
    """
    if vol_range is None:
        vol_range = [0.20, 0.40, 0.60]
    if rate_range is None:
        rate_range = [0.00, 0.025, 0.05]
    
    results = []
    
    for vol in vol_range:
        row = {'Volatility': f"{vol:.0%}"}
        
        for rate in rate_range:
            if method == 'binomial':
                tree = BinomialTree(S0=S0, K=K, T=T, r=rate, sigma=vol, N=N, payoff_type=payoff_type)
                price = tree.price()
            else:
                mc = MonteCarloSimulator(S0=S0, K=K, T=T, r=rate, sigma=vol, payoff_type=payoff_type)
                price = mc.price()
            
            row[f"r={rate:.1%}"] = price
        
        results.append(row)
    
    df = pd.DataFrame(results).set_index('Volatility')
    return df


def export_to_excel(filename: str, 
                   sensitivity_df: Optional[pd.DataFrame] = None,
                   volatility_stress_df: Optional[pd.DataFrame] = None,
                   rate_stress_df: Optional[pd.DataFrame] = None,
                   combined_stress_df: Optional[pd.DataFrame] = None) -> None:
    """
    Export analysis results to Excel workbook.
    
    Parameters
    ----------
    filename : str
        Output filename (e.g., 'analysis.xlsx')
    sensitivity_df : pd.DataFrame, optional
        Sensitivity table
    volatility_stress_df : pd.DataFrame, optional
        Volatility stress test
    rate_stress_df : pd.DataFrame, optional
        Rate stress test
    combined_stress_df : pd.DataFrame, optional
        Combined 2D stress test
    
    Raises
    ------
    ImportError
        If openpyxl not installed
    
    Example
    -------
    >>> sensitivity_df = sensitivity_table(S0=0.035, K=0.040, T=1.0, r=0.025, sigma=0.42)
    >>> vol_stress = stress_test_volatility(S0=0.035, K=0.040, T=1.0, r=0.025)
    >>> export_to_excel('analysis.xlsx', 
    ...                 sensitivity_df=sensitivity_df,
    ...                 volatility_stress_df=vol_stress)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write sheets
        if sensitivity_df is not None:
            sensitivity_df.to_excel(writer, sheet_name='Sensitivity', index=False)
        
        if volatility_stress_df is not None:
            volatility_stress_df.to_excel(writer, sheet_name='Vol Stress', index=False)
        
        if rate_stress_df is not None:
            rate_stress_df.to_excel(writer, sheet_name='Rate Stress', index=False)
        
        if combined_stress_df is not None:
            combined_stress_df.to_excel(writer, sheet_name='Combined Stress')
        
        # Format sheets
        workbook = writer.book
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Header formatting
            for cell in ws[1]:
                if cell.value is not None:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column width auto-adjust
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save
        workbook.save(filename)
    
    print(f"✅ Analysis exported to {filename}")


# Convenience function: run all analyses
def run_full_analysis(S0: float, K: float, T: float, r: float, sigma: float,
                      location: str = "Location",
                      export_file: Optional[str] = None) -> Dict:
    """
    Run comprehensive sensitivity and stress test analysis.
    
    Parameters
    ----------
    S0 : float
        Spot price
    K : float
        Strike price
    T : float
        Time to maturity
    r : float
        Risk-free rate
    sigma : float
        Volatility
    location : str
        Location name for context (default: "Location")
    export_file : str, optional
        Excel filename to export results
    
    Returns
    -------
    Dict
        Contains: sensitivity_table, vol_stress, rate_stress, combined_stress DataFrames
    
    Example
    -------
    >>> results = run_full_analysis(
    ...     S0=0.035, K=0.040, T=1.0, r=0.025, sigma=0.42,
    ...     location='Taiwan',
    ...     export_file='taiwan_analysis.xlsx'
    ... )
    >>> print(results['sensitivity_table'])
    """
    print(f"Running full analysis for {location}...")
    
    sensitivity_df = sensitivity_table(S0=S0, K=K, T=T, r=r, sigma=sigma)
    vol_stress_df = stress_test_volatility(S0=S0, K=K, T=T, r=r)
    rate_stress_df = stress_test_rates(S0=S0, K=K, T=T, sigma=sigma)
    combined_stress_df = combined_stress_test(S0=S0, K=K, T=T)
    
    if export_file:
        export_to_excel(export_file,
                       sensitivity_df=sensitivity_df,
                       volatility_stress_df=vol_stress_df,
                       rate_stress_df=rate_stress_df,
                       combined_stress_df=combined_stress_df)
    
    print("✅ Analysis complete")
    
    return {
        'sensitivity_table': sensitivity_df,
        'vol_stress': vol_stress_df,
        'rate_stress': rate_stress_df,
        'combined_stress': combined_stress_df,
    }


# ============================================================================
# ADDITIONAL QUICK WINS
# ============================================================================

def scenario_comparison(scenarios: List[Dict[str, float]], T: float = 1.0, 
                       r: float = 0.025, method: str = 'binomial', 
                       N: int = 100) -> pd.DataFrame:
    """
    Compare pricing across multiple scenarios side-by-side.
    
    Parameters
    ----------
    scenarios : List[Dict]
        List of scenario dicts: {'name': str, 'S0': float, 'K': float, 'sigma': float}
    T : float
        Time to maturity (default: 1.0 year)
    r : float
        Risk-free rate (default: 2.5%)
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Binomial steps (default: 100)
    
    Returns
    -------
    pd.DataFrame
        Columns: Scenario Name, Spot, Strike, Vol, Price, Delta, Vega, Gamma
    
    Example
    -------
    >>> scenarios = [
    ...     {'name': 'Bullish', 'S0': 0.040, 'K': 0.035, 'sigma': 0.30},
    ...     {'name': 'Base', 'S0': 0.035, 'K': 0.040, 'sigma': 0.42},
    ...     {'name': 'Bearish', 'S0': 0.030, 'K': 0.045, 'sigma': 0.60},
    ... ]
    >>> comp = scenario_comparison(scenarios, T=1.0, r=0.025)
    >>> print(comp)
    """
    results = []
    
    for scenario in scenarios:
        name = scenario.get('name', 'Unnamed')
        S0 = scenario['S0']
        K = scenario['K']
        sigma = scenario['sigma']
        
        # Price
        if method == 'binomial':
            tree = BinomialTree(S0=S0, K=K, T=T, r=r, sigma=sigma, N=N, payoff_type='call')
            price = tree.price()
        else:
            mc = MonteCarloSimulator(S0=S0, K=K, T=T, r=r, sigma=sigma, payoff_type='call')
            price = mc.price()
        
        # Greeks
        greeks = GreeksCalculator(S0=S0, K=K, T=T, r=r, sigma=sigma, pricing_method=method)
        delta = greeks.delta()
        vega = greeks.vega()
        gamma = greeks.gamma()
        
        results.append({
            'Scenario': name,
            'Spot ($/kWh)': S0,
            'Strike ($/kWh)': K,
            'Volatility': f"{sigma:.0%}",
            'Price ($/kWh)': price,
            'Delta': delta,
            'Vega': vega,
            'Gamma': gamma,
        })
    
    df = pd.DataFrame(results)
    return df


def break_even_analysis(S0: float, K: float, T: float, r: float, sigma: float,
                        option_price: Optional[float] = None,
                        method: str = 'binomial', N: int = 100) -> Dict:
    """
    Calculate break-even spot price for option buyer.
    
    Parameters
    ----------
    S0 : float
        Current spot price
    K : float
        Strike price
    T : float
        Time to maturity
    r : float
        Risk-free rate
    sigma : float
        Volatility
    option_price : float, optional
        Premium paid (default: computed)
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Binomial steps
    
    Returns
    -------
    Dict
        Contains: break_even_spot, profit_at_current, max_loss, intrinsic_value
    
    Example
    -------
    >>> be = break_even_analysis(S0=0.035, K=0.040, T=1.0, r=0.025, sigma=0.42)
    >>> print(f"Break-even: ${be['break_even_spot']:.6f}/kWh")
    >>> print(f"Max loss: ${be['max_loss']:.6f}/kWh")
    """
    # Compute option price if not provided
    if option_price is None:
        if method == 'binomial':
            tree = BinomialTree(S0=S0, K=K, T=T, r=r, sigma=sigma, N=N, payoff_type='call')
            option_price = tree.price()
        else:
            mc = MonteCarloSimulator(S0=S0, K=K, T=T, r=r, sigma=sigma, payoff_type='call')
            option_price = mc.price()
    
    # Break-even: K + premium
    break_even = K + option_price
    
    # Current P&L
    intrinsic = max(S0 - K, 0)
    current_pnl = intrinsic - option_price
    
    # Max loss = premium paid (if S0 never rises above K)
    max_loss = -option_price
    
    # Max profit = unlimited (call)
    max_profit = float('inf')
    
    return {
        'break_even_spot': break_even,
        'current_spot': S0,
        'strike': K,
        'option_premium': option_price,
        'current_profit_loss': current_pnl,
        'intrinsic_value': intrinsic,
        'max_loss': max_loss,
        'max_profit': max_profit,
        'profit_margin': (current_pnl / option_price * 100) if option_price > 0 else 0,
    }


def portfolio_greeks(contracts: List[Dict[str, float]], 
                    method: str = 'binomial', N: int = 100) -> Dict[str, float]:
    """
    Aggregate Greeks across a portfolio of contracts.
    
    Parameters
    ----------
    contracts : List[Dict]
        List of contract dicts:
        {'S0': float, 'K': float, 'T': float, 'r': float, 'sigma': float, 'quantity': int}
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Binomial steps
    
    Returns
    -------
    Dict
        Portfolio-level Greeks: delta, gamma, vega, theta, rho
    
    Example
    -------
    >>> contracts = [
    ...     {'S0': 0.035, 'K': 0.040, 'T': 1.0, 'r': 0.025, 'sigma': 0.42, 'quantity': 100},
    ...     {'S0': 0.040, 'K': 0.045, 'T': 0.5, 'r': 0.025, 'sigma': 0.35, 'quantity': 50},
    ... ]
    >>> port_greeks = portfolio_greeks(contracts)
    >>> print(f"Portfolio Delta: {port_greeks['delta']:.2f}")
    >>> print(f"Portfolio Gamma: {port_greeks['gamma']:.2f}")
    """
    portfolio = {
        'delta': 0.0,
        'gamma': 0.0,
        'vega': 0.0,
        'theta': 0.0,
        'rho': 0.0,
    }
    
    for contract in contracts:
        S0 = contract['S0']
        K = contract['K']
        T = contract['T']
        r = contract['r']
        sigma = contract['sigma']
        qty = contract.get('quantity', 1)
        
        greeks = GreeksCalculator(S0=S0, K=K, T=T, r=r, sigma=sigma, pricing_method=method)
        
        portfolio['delta'] += greeks.delta() * qty
        portfolio['gamma'] += greeks.gamma() * qty
        portfolio['vega'] += greeks.vega() * qty
        portfolio['theta'] += greeks.theta() * qty
        portfolio['rho'] += greeks.rho() * qty
    
    return portfolio


def pnl_calculator(S0: float, K: float, T: float, r: float, sigma: float,
                   spot_range: Optional[List[float]] = None,
                   option_price: Optional[float] = None,
                   position: str = 'long_call',
                   method: str = 'binomial', N: int = 100) -> pd.DataFrame:
    """
    Calculate profit/loss at different spot prices at expiry.
    
    Parameters
    ----------
    S0 : float
        Current spot price
    K : float
        Strike price
    T : float
        Time to maturity (for pricing, not used for expiry P&L)
    r : float
        Risk-free rate
    sigma : float
        Volatility (for pricing current option value)
    spot_range : List[float], optional
        Range of spot prices at expiry (default: -30% to +30%)
    option_price : float, optional
        Premium paid (default: computed)
    position : str
        'long_call', 'short_call', 'long_put', 'short_put'
    method : str
        'binomial' or 'monte_carlo'
    N : int
        Binomial steps
    
    Returns
    -------
    pd.DataFrame
        Columns: Spot at Expiry, Intrinsic Value, Premium Cost, Net P&L, Profit/Loss
    
    Example
    -------
    >>> pnl = pnl_calculator(S0=0.035, K=0.040, T=1.0, r=0.025, sigma=0.42,
    ...                      position='long_call')
    >>> print(pnl[['Spot at Expiry', 'Net P&L', 'Profit/Loss']])
    """
    if spot_range is None:
        spot_range = np.linspace(S0 * 0.7, S0 * 1.3, 21)
    else:
        spot_range = np.array(spot_range)
    
    # Compute option price if not provided
    if option_price is None:
        if method == 'binomial':
            tree = BinomialTree(S0=S0, K=K, T=T, r=r, sigma=sigma, N=N, payoff_type='call')
            option_price = tree.price()
        else:
            mc = MonteCarloSimulator(S0=S0, K=K, T=T, r=r, sigma=sigma, payoff_type='call')
            option_price = mc.price()
    
    results = []
    
    for spot_t in spot_range:
        # Intrinsic values at expiry
        call_intrinsic = max(spot_t - K, 0)
        put_intrinsic = max(K - spot_t, 0)
        
        # P&L based on position
        if position == 'long_call':
            pnl = call_intrinsic - option_price
        elif position == 'short_call':
            pnl = option_price - call_intrinsic
        elif position == 'long_put':
            pnl = put_intrinsic - option_price
        elif position == 'short_put':
            pnl = option_price - put_intrinsic
        else:
            raise ValueError(f"Unknown position: {position}")
        
        status = "Profit" if pnl > 0 else ("Loss" if pnl < 0 else "Breakeven")
        
        results.append({
            'Spot at Expiry ($/kWh)': spot_t,
            'Call Intrinsic ($/kWh)': call_intrinsic,
            'Put Intrinsic ($/kWh)': put_intrinsic,
            'Premium ($/kWh)': option_price,
            'Net P&L ($/kWh)': pnl,
            'Profit/Loss': status,
        })
    
    df = pd.DataFrame(results)
    # Simplify columns for display
    if position.startswith('long_call') or position.startswith('short_call'):
        df = df[['Spot at Expiry ($/kWh)', 'Call Intrinsic ($/kWh)', 'Premium ($/kWh)', 'Net P&L ($/kWh)', 'Profit/Loss']]
    else:
        df = df[['Spot at Expiry ($/kWh)', 'Put Intrinsic ($/kWh)', 'Premium ($/kWh)', 'Net P&L ($/kWh)', 'Profit/Loss']]
    
    return df
